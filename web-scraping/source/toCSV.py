#-------------------------------#
# Import libraries              #
#-------------------------------#
import os
import csv
from itertools import islice
from openpyxl import load_workbook

def excel_to_csv(file_path):
    """
    Function that converts the obtained Excel dataset 
    into a CSV file.

    Parameters:
        - file_path (str): path to the Excel file to be converted
    """
    if not os.path.exists(file_path):
        print("Error: File is not found.")
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    csv_path = os.path.splitext(file_path)[0] + ".csv"
    with open(csv_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        for row in islice(sheet.iter_rows(values_only=True), 8, None):
            writer.writerow(row[1:])
